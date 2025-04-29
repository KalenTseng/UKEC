import pandas as pd
import os
import io
import tempfile
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
from pypinyin import pinyin, Style
import re
import time
import zipfile
from openpyxl.worksheet.datavalidation import DataValidation


def clean_excel_value(value):
    """清理Excel中的特殊字符"""
    if not isinstance(value, str):
        return value
    # 替换中文括号
    value = value.replace('（', '(').replace('）', ')')
    # 替换其他可能导致问题的字符
    value = value.replace('\n', ' ').replace('\r', ' ')
    # 移除其他不可见字符
    value = ''.join(char for char in value if ord(char) >= 32)
    # 限制字符串长度
    if len(value) > 32767:  # Excel单元格的最大长度
        value = value[:32767]
    return value

def clean_sheet_name(name):
    """清理工作表名称，移除或替换不允许在Excel工作表名称中使用的字符"""
    if not isinstance(name, str):
        name = str(name)
    # 替换不允许的字符
    cleaned_name = re.sub(r'[\[\]:*?/\\]', '_', name)
    # 移除其他可能导致问题的字符
    cleaned_name = clean_excel_value(cleaned_name)
    # 限制长度为31个字符（Excel的限制）
    if len(cleaned_name) > 31:
        cleaned_name = cleaned_name[:31]
    return cleaned_name

def extract_id_from_hyperlink(value):
    """从超链接公式中提取ID，如果不是超链接则直接返回原值"""
    if isinstance(value, str) and value.startswith('=HYPERLINK'):
        try:
            # 提取引号中的第二个参数（显示文本）
            id_str = value.split('"')[3]
            return int(id_str)
        except (IndexError, ValueError):
            return value
    return value

def extract_name_from_hyperlink(value):
    """从超链接公式中提取姓名，如果不是超链接则直接返回原值"""
    if isinstance(value, str) and value.startswith('=HYPERLINK'):
        try:
            # 提取引号中的最后一个参数（显示文本）
            name = value.split('"')[-2]
            return name
        except (IndexError, ValueError):
            return value
    return value

def get_first_letter(name):
    """获取姓名拼音的首字母"""
    if not name:
        return 'Z'  # 默认返回Z
    # 获取第一个字的拼音首字母
    result = pinyin(name[0], style=Style.FIRST_LETTER)
    if result and result[0]:
        return result[0][0].upper()
    return 'Z'  # 如果无法获取拼音，返回Z

def save_workbook_with_retry(workbook, max_retries=3, delay=2):
    """保存工作簿到内存中并返回字节流"""
    output = io.BytesIO()
    for attempt in range(max_retries):
        try:
            workbook.save(output)
            output.seek(0)
            return output
        except Exception as e:
            if attempt < max_retries - 1:
                st.warning(f"保存失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}，{delay}秒后重试...")
                time.sleep(delay)
            else:
                st.error(f"保存失败: {str(e)}")
                raise e

def create_offer_sheets(template_file, kalen_file, selected_sheet):
    st.info("开始处理数据...")
    
    try:
        # 读取模板文件
        st.info("正在读取模板文件...")
        template_df = pd.read_csv(template_file, encoding='utf-8')
        st.success("成功读取模板文件")
        
        # 清理列名中的换行符
        template_df.columns = [col.strip() for col in template_df.columns]
        
        st.info("正在处理数据列...")
        # 定义需要保留的列
        columns_to_keep = [
            '客户id', '姓名',  # 这两列用于识别，但不会显示在最终表格中
            '申请院校英文', '申请专业英文', '申请结果', 
            'Student-Number',  '入学条件',
            '是否完成语言条件', '语言条件', '是否完成推荐信条件', '推荐信条件',
            '押金类型', '押金截止日期', '是否完成押金条件', '成押金条件'
        ]
        
        # 检查哪些列存在，哪些列不存在
        existing_columns = [col for col in columns_to_keep if col in template_df.columns]
        missing_columns = [col for col in columns_to_keep if col not in template_df.columns]
        
        st.info(f"找到 {len(existing_columns)} 个有效列")
        if missing_columns:
            st.warning(f"注意：以下列不存在：{', '.join(missing_columns)}")
        
        # 只保留存在的列
        template_df = template_df[existing_columns]
        st.success("数据列处理完成")
        
        st.info("正在处理日期数据...")
        # 将"未获得"替换为空值
        template_df = template_df.replace('未获得', '')
        
        # 尝试将押金截止日期列转换为日期格式
        def parse_date(date_str):
            if pd.isna(date_str):
                return None
            try:
                if isinstance(date_str, str):
                    if ':' in date_str:  # 如果包含时间
                        return pd.to_datetime(date_str.split()[0])  # 只取日期部分
                    return pd.to_datetime(date_str)
                return date_str
            except:
                return None
        
        template_df['押金截止日期'] = template_df['押金截止日期'].apply(parse_date)
        st.success("日期处理完成")
        
        # 从申请跟进表中读取数据
        st.info("正在读取申请跟进表...")
        # 使用临时文件来处理Excel文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(kalen_file.getvalue())
            tmp_path = tmp.name
        
        kalen_wb = load_workbook(tmp_path, data_only=True)
        kalen_ws = kalen_wb[selected_sheet]
        st.success("成功读取申请跟进表")
        
        # 找到ID列
        id_col = None
        for col in range(1, kalen_ws.max_column + 1):
            if kalen_ws.cell(row=1, column=col).value == '客户id':
                id_col = col
                break
        
        st.info("正在提取ID信息...")
        # 读取ID列的值（保存原始ID值）
        id_values = {}
        if id_col:
            for row in range(2, kalen_ws.max_row + 1):
                cell = kalen_ws.cell(row=row, column=id_col)
                original_value = cell.value
                # 如果是超链接公式，提取ID值
                clean_id = extract_id_from_hyperlink(original_value)
                if clean_id:
                    id_values[row] = clean_id
        st.success(f"成功提取 {len(id_values)} 个ID")
        
        # 关闭工作簿
        kalen_wb.close()
        
        # 创建新的Offer跟进表 - 使用Workbook()创建空工作簿而不是load_workbook()
        st.info("开始创建Offer跟进表...")
        offer_wb = Workbook()
        
        # 获取当前日期
        today = datetime.now().date()
        
        # 定义要显示的列（不包括客户id和姓名）
        display_columns = [col for col in existing_columns if col not in ['客户id', '姓名']]
        
        # 获取总客户数
        total_clients = len(template_df['客户id'].unique())
        st.info(f"共有 {total_clients} 个客户需要处理")
        
        # 创建进度条
        progress_bar = st.progress(0)
        
        # 为每个客户ID创建表格
        for idx, client_id in enumerate(template_df['客户id'].unique(), 1):
            # 更新进度条
            progress = idx / total_clients
            progress_bar.progress(progress)
            
            # 获取该客户的所有信息
            client_data = template_df[template_df['客户id'] == client_id].copy()
            
            # 定义申请结果的排序优先级
            result_priority = {
                '获得CAS/COE': 1,
                '获得UO': 2,
                '获得CO': 3,
                '拒信': 4,
                '大学撤销': 5
            }
            
            # 添加排序键列
            client_data.loc[:, 'sort_key'] = client_data['申请结果'].map(lambda x: result_priority.get(x, 6))
            
            # 按排序键和申请结果排序
            client_data = client_data.sort_values(['sort_key', '申请结果'])
            
            # 删除排序键列
            client_data = client_data.drop('sort_key', axis=1)
            
            # 获取客户姓名
            if len(client_data) > 0 and '姓名' in client_data.columns:
                client_name = client_data['姓名'].iloc[0] if not pd.isna(client_data['姓名'].iloc[0]) else "未知姓名"
            else:
                client_name = "未知姓名"
            
            # 如果没有数据，跳过此客户
            if len(client_data) == 0:
                continue
            
            # 如果表格已存在，则删除
            if clean_sheet_name(str(client_id)) in offer_wb.sheetnames:
                offer_wb.remove(offer_wb[clean_sheet_name(str(client_id))])
            
            # 创建新的表格
            ws = offer_wb.create_sheet(title=clean_sheet_name(str(client_id)))
            
            # 添加姓名行（带超链接）
            name_cell = ws.cell(row=1, column=1)
            crm_link = f'http://crmuk.ukec.com/admin/school-apply/detail?id={client_id}'
            name_cell.value = f'=HYPERLINK("{crm_link}", "{client_name}")'
            name_cell.font = Font(bold=True, size=12, color="0000FF", underline="single")  # 蓝色下划线
            name_cell.alignment = Alignment(horizontal='left')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_columns))
            
            # 添加邮箱超链接到N1单元格
            email_cell = ws.cell(row=1, column=14)  # N列是第14列
            email_link = f'http://crmuk.ukec.com/admin/customer-email/detail?id={client_id}'
            email_cell.value = f'=HYPERLINK("{email_link}", "邮箱")'
            email_cell.font = Font(bold=True, size=12, color="0000FF", underline="single")  # 蓝色下划线
            email_cell.alignment = Alignment(horizontal='left')
            
            # 设置表头
            headers = display_columns + ['押金截止倒计时(天)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=header)
                # 设置表头背景色为浅灰色
                ws.cell(row=2, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # 填充数据
            for row_idx, row in enumerate(client_data.iterrows(), 3):  # 从第3行开始，因为第1行是姓名，第2行是表头
                # 填充原有数据（跳过客户id和姓名）
                col_idx = 1  # 从第1列开始
                for field in display_columns:
                    value = row[1][field]  # 使用字典方式访问，避免特殊字符问题
                    # 清理值中的特殊字符
                    value = clean_excel_value(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 如果是"未获得"，则设置为空值
                    if value == '未获得':
                        cell.value = ''
                    
                    # 为申请结果列添加条件格式
                    if field == '申请结果' and value:
                        # 绿色填充，白色文字：获得CO，获得UO，获得CAS/COE
                        if '获得' in str(value):
                            cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # 绿色
                            cell.font = Font(color='FFFFFF', bold=True)  # 白色文字，加粗
                        # 灰色填充：拒信，大学撤销
                        elif value in ['拒信', '大学撤销','学生放弃']:
                            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # 灰色
                            cell.font = Font(bold=True)  # 加粗
                        # 橙色填充：已递交
                        elif value in ['已递交']:
                            cell.fill = PatternFill(start_color='FFC222', end_color='FFC222', fill_type='solid')  # 橙色
                            cell.font = Font(bold=True)  # 加粗
                        # 橙色填充：未递-待交申请费
                        elif value in ['未递-待交申请费','已递交-申请费待付']:
                            cell.fill = PatternFill(start_color='5F9EA0', end_color='5F9EA0', fill_type='solid')  # 蓝色
                            cell.font = Font(bold=True)  # 加粗
                        # 橙色填充：未递-材料缺失或需修改
                        elif value in ['未递-材料缺失或需修改']:
                            cell.fill = PatternFill(start_color='F08080', end_color='F08080', fill_type='solid')  # 橙色
                            cell.font = Font(bold=True)  # 加粗
                    
                    col_idx += 1
                
                # 计算押金截止日期倒计时
                deadline = row[1]['押金截止日期']
                
                if pd.notna(deadline):
                    try:
                        # 计算剩余天数
                        days_left = (deadline.date() - today).days
                        
                        # 添加倒计时列
                        countdown_cell = ws.cell(row=row_idx, column=len(headers), value=days_left)
                        
                        # 设置单元格格式
                        if days_left <= 30 and days_left > 0:
                            # 30天内红色背景白字
                            countdown_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            countdown_cell.font = Font(color='FFFFFF')
                        elif days_left <= 0:
                            # 已过期灰色背景
                            countdown_cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                    except (ValueError, AttributeError):
                        pass
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 删除默认的Sheet
        if 'Sheet' in offer_wb.sheetnames:
            offer_wb.remove(offer_wb['Sheet'])
        
        # 更新申请跟进.xlsx中的超链接，仅添加ID列超链接，不添加姓名列文件夹链接
        st.info("更新申请跟进表中的超链接...")
        kalen_wb = load_workbook(tmp_path)
        kalen_ws = kalen_wb[selected_sheet]
        
        # 在ID列添加超链接，使用保存的原始ID值
        if id_col:
            for row, original_id in id_values.items():
                if original_id:
                    # 创建超链接
                    link = f'=HYPERLINK("[Offer 跟进.xlsx]{original_id}!A1", "{original_id}")'
                    kalen_ws.cell(row=row, column=id_col, value=link)
        
        # 创建或更新VIP情况表
        st.info("更新VIP情况表...")
        if "VIP情况" in kalen_wb.sheetnames:
            kalen_wb.remove(kalen_wb["VIP情况"])
        vip_ws = kalen_wb.create_sheet(title="VIP情况", index=0)  # 设置为第一个表
        
        # 定义VIP学校列表
        vip_schools = [
            "University College London",
            "The London School of Economics and Political Science",
            "King's College London",
            "University of Oxford",
            "University of Cambridge"
        ]
        
        # 设置表头
        vip_headers = ['姓名', '申请院校英文', '申请专业英文', '申请结果']
        for col, header in enumerate(vip_headers, 1):
            vip_ws.cell(row=1, column=col, value=header)
            # 设置表头背景色为浅灰色
            vip_ws.cell(row=1, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 收集VIP学校的申请数据
        vip_data = []
        for sheet_name in offer_wb.sheetnames:
            if sheet_name in ["押金DDL", "VIP情况"]:
                continue
                
            ws = offer_wb[sheet_name]
            client_name = extract_name_from_hyperlink(ws.cell(row=1, column=1).value)
            client_id = sheet_name  # 使用sheet名称作为客户ID
            
            # 从第3行开始读取数据（跳过姓名行和表头）
            for row in range(3, ws.max_row + 1):
                school = ws.cell(row=row, column=1).value
                if any(vip_school.lower() in str(school).lower() for vip_school in vip_schools):
                    # 获取其他列的数据
                    program = ws.cell(row=row, column=2).value
                    result = ws.cell(row=row, column=3).value
                    
                    vip_data.append({
                        '姓名': client_name,
                        '客户id': client_id,  # 添加客户ID但不显示在表格中
                        '申请院校英文': school,
                        '申请专业英文': program,
                        '申请结果': result
                    })
        
        # 按学校名称和申请结果排序
        vip_data.sort(key=lambda x: (
            x['申请院校英文'] or '',
            0 if '获得' in str(x['申请结果'] or '') else 1,  # 获得offer的排在前面
            x['姓名'] or ''
        ))
        
        # 填充数据
        for row_idx, data in enumerate(vip_data, 2):
            for col_idx, header in enumerate(vip_headers, 1):
                if header == '姓名':
                    # 创建CRM链接
                    crm_link = f'http://crmuk.ukec.com/admin/school-apply/detail?id={data["客户id"]}'
                    cell = vip_ws.cell(row=row_idx, column=col_idx)
                    cell.value = f'=HYPERLINK("{crm_link}", "{data[header]}")'
                    cell.font = Font(color="0000FF", underline="single")  # 蓝色下划线
                else:
                    cell = vip_ws.cell(row=row_idx, column=col_idx, value=data[header])
                
                # 为申请结果添加条件格式
                if header == '申请结果':
                    result = data[header]
                    if result and '获得' in str(result):
                        cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # 绿色
                        cell.font = Font(color='FFFFFF', bold=True)  # 白色文字，加粗
                    elif result in ['拒信', '大学撤销','未递-撤销申请需求']:
                        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # 灰色
                        cell.font = Font(bold=True)  # 加粗
                    elif result in ['已递交']:
                        cell.fill = PatternFill(start_color='FFC222', end_color='FFC222', fill_type='solid')  # 橙色
                        cell.font = Font(bold=True)  # 加粗
        
        # 调整列宽
        for col in range(1, len(vip_headers) + 1):
            vip_ws.column_dimensions[get_column_letter(col)].width = 25
        
        st.success("VIP情况表更新完成")
        
        # 创建或更新押金DDL总结表格
        st.info("更新押金DDL总结表格...")
        if "押金DDL" in kalen_wb.sheetnames:
            kalen_wb.remove(kalen_wb["押金DDL"])
        summary_ws = kalen_wb.create_sheet(title="押金DDL", index=1)  # 设置为第二个表
        
        # 设置表头
        headers = ['姓名', '申请院校英文', '申请专业英文', '押金截止日期', '剩余天数']
        for col, header in enumerate(headers, 1):
            summary_ws.cell(row=1, column=col, value=header)
            # 设置表头背景色为浅灰色
            summary_ws.cell(row=1, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 收集所有近一个月的押金截止日期数据
        summary_data = []
        for sheet_name in offer_wb.sheetnames:
            if sheet_name in ["押金DDL", "VIP情况"]:
                continue
                
            ws = offer_wb[sheet_name]
            client_name = extract_name_from_hyperlink(ws.cell(row=1, column=1).value)
            client_id = sheet_name  # 使用sheet名称作为客户ID
            
            # 从第3行开始读取数据（跳过姓名行和表头）
            for row in range(3, ws.max_row + 1):
                deadline = ws.cell(row=row, column=ws.max_column).value
                if isinstance(deadline, (int, float)) and 0 <= deadline <= 30:  # 只收集30天内的数据
                    # 获取其他列的数据
                    school = ws.cell(row=row, column=1).value
                    program = ws.cell(row=row, column=2).value
                    summary_data.append({
                        '姓名': client_name,
                        '客户id': client_id,  # 添加客户ID但不显示在表格中
                        '申请院校英文': school,
                        '申请专业英文': program,
                        '押金截止日期': deadline,
                        '剩余天数': deadline
                    })
        
        # 按剩余天数排序
        summary_data.sort(key=lambda x: x['剩余天数'])
        
        # 填充数据
        for row_idx, data in enumerate(summary_data, 2):
            for col_idx, header in enumerate(headers, 1):
                if header == '姓名':
                    # 创建CRM链接
                    crm_link = f'http://crmuk.ukec.com/admin/school-apply/detail?id={data["客户id"]}'
                    cell = summary_ws.cell(row=row_idx, column=col_idx)
                    cell.value = f'=HYPERLINK("{crm_link}", "{data[header]}")'
                    cell.font = Font(color="0000FF", underline="single")  # 蓝色下划线
                else:
                    cell = summary_ws.cell(row=row_idx, column=col_idx, value=data[header])
                
                # 为剩余天数添加条件格式
                if header == '剩余天数':
                    days = data['剩余天数']
                    if days <= 7:
                        # 7天内红色背景白字
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFF')
                    elif days <= 14:
                        # 8-14天黄色背景
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    elif days <= 30:
                        # 15-30天绿色背景
                        cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                        cell.font = Font(color='FFFFFF')
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            summary_ws.column_dimensions[get_column_letter(col)].width = 20
        
        st.success("押金DDL总结表格更新完成")
        
        # 添加入学院校和入学专业列到申请跟进表
        st.info("添加入学院校和入学专业列...")
        # 找到最后一列
        last_col = kalen_ws.max_column
        
        # 添加入学院校列
        enrollment_school_col = last_col + 1
        kalen_ws.cell(row=1, column=enrollment_school_col, value='入学院校')
        kalen_ws.cell(row=1, column=enrollment_school_col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 添加入学专业列
        enrollment_program_col = last_col + 2
        kalen_ws.cell(row=1, column=enrollment_program_col, value='入学专业')
        kalen_ws.cell(row=1, column=enrollment_program_col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 调整新列的列宽
        kalen_ws.column_dimensions[get_column_letter(enrollment_school_col)].width = 25
        kalen_ws.column_dimensions[get_column_letter(enrollment_program_col)].width = 25
        
        # 创建选项表（辅助表）
        if "选项列表" in kalen_wb.sheetnames:
            kalen_wb.remove(kalen_wb["选项列表"])
        options_ws = kalen_wb.create_sheet(title="选项列表", index=len(kalen_wb.sheetnames))
        options_ws.sheet_state = 'hidden'  # 隐藏此工作表
        
        # 为每个学生收集申请的学校和专业
        student_options = {}
        
        # 遍历申请跟进表中的每一行，找到ID
        for row in range(2, kalen_ws.max_row + 1):
            id_cell = kalen_ws.cell(row=row, column=id_col)
            if id_cell.value:
                # 提取ID（可能是超链接）
                student_id = extract_id_from_hyperlink(id_cell.value)
                if student_id:
                    student_options[row] = {
                        'id': student_id,
                        'schools': set(),
                        'programs': set()
                    }
        
        # 遍历Offer 跟进.xlsx中的每个工作表（客户ID）
        for sheet_name in offer_wb.sheetnames:
            if sheet_name in ["押金DDL", "VIP情况"]:
                continue
                
            try:
                student_id = int(sheet_name)  # 尝试将工作表名转换为ID
                ws = offer_wb[sheet_name]
                
                # 从第3行开始读取数据（跳过名称和表头）
                for row in range(3, ws.max_row + 1):
                    school = ws.cell(row=row, column=1).value  # 申请院校英文 (第一列)
                    program = ws.cell(row=row, column=2).value  # 申请专业英文 (第二列)
                    
                    if school and program:
                        # 清理值
                        school = clean_excel_value(str(school).strip())
                        program = clean_excel_value(str(program).strip())
                        
                        # 为相应的学生添加学校和专业选项
                        for kalen_row, options in student_options.items():
                            if options['id'] == student_id:
                                options['schools'].add(school)
                                options['programs'].add(program)
                                break
            except (ValueError, KeyError):
                # 不是有效的客户ID或找不到工作表，跳过
                continue
        
        # 在辅助表中创建每个学生的选项列表
        current_row = 1
        
        for row, options in student_options.items():
            if not options['schools'] and not options['programs']:
                continue
                
            try:
                # 记录该学生的选项起始行
                options_start_row = current_row
                
                # 写入学校选项
                max_count = max(len(options['schools']), len(options['programs']), 1)
                
                # 写入标题
                options_ws.cell(row=current_row, column=1, value=f"ID_{options['id']}_学校")
                options_ws.cell(row=current_row, column=2, value=f"ID_{options['id']}_专业")
                current_row += 1
                
                # 初始化计数器
                schools_written = 0
                programs_written = 0
                
                # 转换为列表，便于索引
                schools_list = list(options['schools'])
                programs_list = list(options['programs'])
                
                # 写入选项数据
                for i in range(max_count):
                    if i < len(schools_list):
                        options_ws.cell(row=current_row, column=1, value=schools_list[i])
                        schools_written += 1
                        
                    if i < len(programs_list):
                        options_ws.cell(row=current_row, column=2, value=programs_list[i])
                        programs_written += 1
                        
                    current_row += 1
                
                # 留一行空白作为分隔
                current_row += 1
                
                # 添加数据验证
                if schools_written > 0:
                    # 创建学校下拉列表
                    school_range = f"选项列表!$A${options_start_row+1}:$A${options_start_row+schools_written}"
                    school_dv = DataValidation(type="list", formula1=f"={school_range}")
                    # 设置错误消息
                    school_dv.error = '请从下拉列表中选择一个选项'
                    school_dv.errorTitle = '输入错误'
                    
                    # 添加到工作表
                    school_cell = f"{get_column_letter(enrollment_school_col)}{row}"
                    school_dv.add(school_cell)
                    kalen_ws.add_data_validation(school_dv)
                
                if programs_written > 0:
                    # 创建专业下拉列表
                    program_range = f"选项列表!$B${options_start_row+1}:$B${options_start_row+programs_written}"
                    program_dv = DataValidation(type="list", formula1=f"={program_range}")
                    # 设置错误消息
                    program_dv.error = '请从下拉列表中选择一个选项'
                    program_dv.errorTitle = '输入错误'
                    
                    # 添加到工作表
                    program_cell = f"{get_column_letter(enrollment_program_col)}{row}"
                    program_dv.add(program_cell)
                    kalen_ws.add_data_validation(program_dv)
            
            except Exception as e:
                st.warning(f"为第{row}行添加下拉选项时出错: {str(e)}")
        
        st.success("入学院校和入学专业列添加完成")
        
        st.success("所有数据处理完成！")
        
        # 保存Offer跟进表到内存中
        offer_output = save_workbook_with_retry(offer_wb)
        
        # 保存更新后的申请跟进表到内存中
        kalen_output = save_workbook_with_retry(kalen_wb)
        
        # 生成一个包含两个文件的ZIP文件
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 将Offer跟进表添加到ZIP文件
            zipf.writestr("Offer 跟进.xlsx", offer_output.getvalue())
            # 将申请跟进表添加到ZIP文件
            zipf.writestr("申请跟进.xlsx", kalen_output.getvalue())
        
        zip_buffer.seek(0)
        
        # 清理临时文件
        try:
            os.unlink(tmp_path)
        except:
            pass
        
        return zip_buffer
        
    except Exception as e:
        st.error(f"处理过程中出现错误: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def get_sheet_names(excel_file):
    """从Excel文件中获取工作表名称列表"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(excel_file.getvalue())
            tmp_path = tmp.name
        
        wb = load_workbook(tmp_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # 清理临时文件
        try:
            os.unlink(tmp_path)
        except:
            pass
            
        return sheet_names
    except Exception as e:
        st.error(f"获取工作表名称时出错: {str(e)}")
        return []

def main():
    st.set_page_config(page_title="Offer跟进表生成器", page_icon="📊", layout="wide")
    
    st.title("Offer跟进表生成器")
    st.markdown("上传文件并生成Offer跟进表和更新后的申请跟进表")
    
    # 文件上传区域
    col1, col2 = st.columns(2)
    
    with col1:
        template_file = st.file_uploader("上传院校申请列表.csv 文件", type=["csv"], accept_multiple_files=False)
    
    with col2:
        kalen_file = st.file_uploader("上传客户列表.XLSX文件", type=["xlsx"], accept_multiple_files=False)
    
    # 如果两个文件都已上传
    if template_file and kalen_file:
        # 获取申请跟进表中的所有工作表
        sheet_names = get_sheet_names(kalen_file)
        
        if sheet_names:
            # 选择工作表
            selected_sheet = st.selectbox("选择要处理的工作表:", sheet_names)
            
            # 生成按钮
            if st.button("生成Offer跟进表"):
                with st.spinner("正在处理数据，请稍候..."):
                    zip_output = create_offer_sheets(template_file, kalen_file, selected_sheet)
                    
                    if zip_output:
                        st.success("处理完成！")
                        
                        # 创建下载按钮
                        st.download_button(
                            label="下载所有生成的文件",
                            data=zip_output,
                            file_name="Offer跟进文件.zip",
                            mime="application/zip"
                        )

if __name__ == '__main__':
    main() 
